import camelot
import pandas as pd
from tkinter import Label, Tk, messagebox
from tkinter.filedialog import askopenfilename
import numpy as np
import shutil
import re
import os
from openpyxl.styles import Border, Side, PatternFill, Alignment, Font
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl import load_workbook

root = Tk()
root.iconify()
w = Label(root, text ='Kiara Jewellery', font = "50")  
w.pack() 
fl = askopenfilename(title="Select File for Order Raising")
tb = camelot.read_pdf(fl, flavor='lattice', pages='1-end')

exl = "Z:\\MARKETING  TEAM\\Transactions\\EMR Import Excel\\ORDER IMPORT EXCEL (P).xlsx"
mpl = "Z:\\MARKETING  TEAM\\Transactions\\Other customer\\Messika(OS39)\\Formats\\For Invoicing\\New MPL 2024\\MPL Master file OS39.xlsx"

conftp = "Z:\\MARKETING  TEAM\\Transactions\\Other customer\\Messika(OS39)\\Formats\\Order Confirmation.xlsx"
dest = "Z:\\MARKETING  TEAM\\Transactions\\Other customer\Messika(OS39)\\PO\\"


def set_border(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

def confirm(df, odr, qty):
    if odr == "EU36":
        messagebox.showinfo("Order Raising", "Import Complete of " + str(len(df)) + " Lines with " + str(qty) + " Quantity") 
    elif odr == "OS39":
        messagebox.showinfo("Order Raising", "Import Complete of " + str(qty) + " Lines with " + str(len(df)) + " Quantity") 

# Order confirmation
def ordconf(dfc):
    dfc[3] = dfc[3].str[-2:]
    dfc[2] = dfc[2] + '-' + dfc[3]
    dfc[2] = dfc[2].str.replace('-NS', '')
    dfc = dfc.drop(dfc.columns[[0, 3, 4, 5, 6, 7]], axis=1)
    dfc.insert(0, 0, dfc.pop(8))

    ml = pd.ExcelFile(mpl)
    exm = pd.read_excel(ml, 'Final Master Cost sheet')
    df = pd.DataFrame()

    for i in range(len(dfc)):
        df = df._append(exm.loc[np.where(exm['Unnamed: 2'].str.contains(dfc.iloc[i][2]) & exm['Unnamed: 1'].str.contains(dfc.iloc[i][1]))])

    df['Unnamed: 0'] = dfc[0].values
    df.reset_index(inplace=True, drop=True)
    df.columns = [i for i in range(len(df.columns))]
    df.fillna('', inplace=True)

    opo = re.findall(r'\d{4,}', os.path.basename(fl))[0]
    opo = dest + opo + ' - Order Confirmation.xlsx' 
    shutil.copy(conftp, opo)

    with pd.ExcelWriter(opo, engine="openpyxl", mode='a', if_sheet_exists="overlay") as writer:
        df.to_excel(writer, startcol=0, startrow=3, sheet_name='Final Master Cost sheet', index=False, header=False)

    wb = load_workbook(filename=opo)
    ws = wb.worksheets[0]
    ws.conditional_formatting = ConditionalFormattingList()
    set_border(ws, "A4:EC" + str((4+len(dfc))))

    for i in range(4, 4+len(dfc)+1):
        for j in range(65, 91):
            ws[chr(j) + str(i)].fill = PatternFill(fgColor="FFFFFF", fill_type = "solid")
            ws[chr(j) + str(i)].alignment = Alignment(horizontal='center')
            ws[chr(j) + str(i)].font = Font(name="Eras Medium ITC", size=10)

            ws['A' + chr(j) + str(i)].fill = PatternFill(fgColor="FFFFFF", fill_type = "solid")
            ws['A' + chr(j) + str(i)].alignment = Alignment(horizontal='center')
            ws['A' + chr(j) + str(i)].font = Font(name="Eras Medium ITC", size=10)

            ws['B' + chr(j) + str(i)].fill = PatternFill(fgColor="FFFFFF", fill_type = "solid")
            ws['B' + chr(j) + str(i)].alignment = Alignment(horizontal='center')
            ws['B' + chr(j) + str(i)].font = Font(name="Eras Medium ITC", size=10)

            ws['C' + chr(j) + str(i)].fill = PatternFill(fgColor="FFFFFF", fill_type = "solid")
            ws['C' + chr(j) + str(i)].alignment = Alignment(horizontal='center')
            ws['C' + chr(j) + str(i)].font = Font(name="Eras Medium ITC", size=10)

            ws['D' + chr(j) + str(i)].fill = PatternFill(fgColor="FFFFFF", fill_type = "solid")
            ws['D' + chr(j) + str(i)].alignment = Alignment(horizontal='center')
            ws['D' + chr(j) + str(i)].font = Font(name="Eras Medium ITC", size=10)
            
        ws['A' + str(i)].font = Font(name="Eras Medium ITC", size=11)

        ws['EA' + str(i)].fill = PatternFill(fgColor="FFFFFF", fill_type = "solid")
        ws['EA' + str(i)].alignment = Alignment(horizontal='center')
        ws['EA' + str(i)].font = Font(name="Eras Medium ITC", size=10)

        ws['EB' + str(i)].fill = PatternFill(fgColor="FFFFFF", fill_type = "solid")
        ws['EB' + str(i)].alignment = Alignment(horizontal='center')
        ws['EB' + str(i)].font = Font(name="Eras Medium ITC", size=10)
        ws['EC' + str(i)].fill = PatternFill(fgColor="FFFFFF", fill_type = "solid")
        ws['EC' + str(i)].alignment = Alignment(horizontal='center')
        ws['EC' + str(i)].font = Font(bold=True, name="Eras Medium ITC", size=10)
        ws['EC' + str(i)].number_format = '"$"#,##0'

        ws['P' + str(i)].font = Font(name="Eras Medium ITC", color="0066CC", bold=True, size=10)
        ws['Q' + str(i)].font = Font(name="Eras Medium ITC", color="0066CC", bold=True, size=10)

    li = ['K', 'L', 'M', 'N', 'O', 'Q', 'T', 'Z', 'AF', 'AL', 'AR', 'AX', 'BD', 'BJ', 'BP', 'BV', 'CB', 'CH', 'CN', 'CT', 'CZ']

    for i in range(4, 4+len(dfc)+1):
        for j in li:
            ws[j + str(i)].number_format = '0.000'


    for i in range(4, 4+len(dfc)+1):
        for j in range(68, 90):
            ws['D' + chr(j) + str(i)].number_format = '"$"#,##0.00'

    for i in range(4, 4+len(dfc)+1):
        ws['P' + str(i)].fill = PatternFill(fgColor="CCCCFF", fill_type = "solid")
        ws['Q' + str(i)].fill = PatternFill(fgColor="CCCCFF", fill_type = "solid")
        ws['DD' + str(i)].fill = PatternFill(fgColor="FFFF99", fill_type = "solid")
        ws['DE' + str(i)].fill = PatternFill(fgColor="FFFF99", fill_type = "solid")
        ws['EC' + str(i)].fill = PatternFill(fgColor="BFBFBF", fill_type = "solid")
        ws['DY' + str(i)].fill = PatternFill(fgColor="CCFFCC", fill_type = "solid")
        ws['DZ' + str(i)].fill = PatternFill(fgColor="FFFF99", fill_type = "solid")
        ws['EB' + str(i)].fill = PatternFill(fgColor="FFFF99", fill_type = "solid")
        ws['EA' + str(i)].fill = PatternFill(fgColor="FFFF99", fill_type = "solid")

        for j in range(70, 85):
            ws['D' + chr(j) + str(i)].fill = PatternFill(fgColor="FFCC00", fill_type = "solid")
        for k in range(85, 89):
            ws['D' + chr(k) + str(i)].fill = PatternFill(fgColor="FFFF99", fill_type = "solid")

    wb.save(opo)
    

# EU36
def lp(df):
    frames = []
    for i in range(len(df)):
        if df.iloc[i][1].isdigit():
            frames.append(df.iloc[i].tolist())
    df = pd.DataFrame(frames)
    df.replace('', None, regex=True, inplace=True)
    df[3].fillna(df[4], inplace=True)
    df.drop([4], axis=1, inplace=True)
    df.update(df[3].str[:10])

    df[7] = np.where(df[2].str.contains('BR', regex=False), "B" + df[7] + "0", df[7])
    df[7] = np.where(df[2].str.contains('NK', regex=False), df[7] + "CM", df[7])
    df[7] = np.where(df[2].str.contains('PD', regex=False), df[7] + "CM", df[7])
    df[7] = np.where(df[2].str.contains('RG', regex=False), "R" + df[7], df[7])
    df[7] = np.where(df[2].str.contains('ER', regex=False), "NS", df[7])

    df.drop([0, 5, 6, 9], axis=1, inplace=True)
    df[1] = df[1].astype(int)
    df[8] = df[8].astype(int)
    df.columns = [i for i in range(len(df.columns))]
    return df

def euquery(df):
    with pd.ExcelWriter(exl, engine="openpyxl", mode='a', if_sheet_exists="overlay") as writer:
        df[0].to_excel(writer, startcol=0, startrow=1, index=False, header=False)
        df.iloc[:, 1:5].to_excel(writer, startcol=2, startrow=1, index=False, header=False)
        df[4].to_excel(writer, startcol=6, startrow=1, index=False, header=False)
    confirm(df, "EU36", df[4].sum())

def eu36():
    df = lp(pd.concat([tb[i].df for i in range(tb.n)])) 
    #dt = [dt[:6] + dt[8:] for dt in re.findall(r'\d{2}/\d{2}/\d{4}', tb[0].df.loc[1][0])]
    #po = re.findall(r'\n\d{6}', tb[0].df.loc[1][0])[0]
    #fr = tb[0].df.loc[1][0].split('\n')[-1].split()[1]
    euquery(df)

# OS39
def increment(dt):
    if len(dt) == 1:
        return [dt[0][0], int(dt[0][1:])]
    else:
        return [dt[1][0], int(dt[1][1:])]

def mes(df):
    frames = []
    sr = 1

    for i in range(1, len(df), 2):
        dt = df.iloc[i][0].splitlines()
        if (len(dt) > 7):
            if (dt[8] == "To engrave"):
                ino = increment(dt[9].split())
                for j in range(0, len(dt)-len(dt)%10, 10):
                    for k in range(int(dt[j+3])):
                        if dt[j+2].count("-") == 1:
                            dt[j+2] += "-NS"
                        frames.append([sr] + [0] + [dt[j+2][:-3]] + [dt[j+2][-2:]] + [1] + [1] + ["MESSIKA Au750 " + ino[0] + str(ino[1])] + [dt[j+1]] + [ino[0] + str(ino[1])])
                        ino[1] += 1
                        sr += 1

            elif (dt[9] == "To engrave"):
                ino = increment(dt[10].split())
                for j in range(0, len(dt)-len(dt)%11, 11):
                    for k in range(int(dt[j+3])):
                        if dt[j+2].count("-") == 1:
                            dt[j+2] += "-NS"
                        frames.append([sr] + [dt[j+8]] + [dt[j+2][:-3]] + [dt[j+2][-2:]] + [1] + [1] + ["MESSIKA Au750 " + ino[0] + str(ino[1])] + [dt[j+1]] + [ino[0] + str(ino[1])])
                        ino[1] += 1
                        sr += 1

    df = pd.DataFrame(frames)
    df.replace('NS', None, regex=True, inplace=True)
    df[3] = "R" + df[3]
    df[3].fillna("NS", inplace=True)
    return df

def os39query(df):
    with pd.ExcelWriter(exl, engine="openpyxl", mode='a', if_sheet_exists="overlay") as writer:
        df[0].to_excel(writer, startcol=0, startrow=1, index=False, header=False)
        df.iloc[:, 1:6].to_excel(writer, startcol=2, startrow=1, index=False, header=False)
        df.iloc[:, 6:8].to_excel(writer, startcol=9, startrow=1, index=False, header=False)
        df[8].to_excel(writer, startcol=22, startrow=1, index=False, header=False)
    ordconf(df)
    confirm(df, "OS39", df[7].nunique())

def os39():
    df = mes(pd.concat([tb[i].df for i in range(tb.n)]))
    os39query(df)

if "LP CREATIONS" in tb[0].df.loc[0][0]:
    eu36()
else:  
    os39()
